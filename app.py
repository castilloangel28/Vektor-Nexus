from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os
import random
import string
import io
import openpyxl  # Para procesar hojas de cálculo Excel (.xlsx) desde Google Drive
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from datetime import datetime

# --- LIBRERÍAS PARA PDF E IMÁGENES ---
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'vektor_nexus_secure_key_2026_angel')

# Identificador de carpeta raíz en Google Drive
CARPETA_RAIZ_DRIVE = os.environ.get('DRIVE_FOLDER_ID', "1PbH8767Q86O-TntoxDxozaGiBl3WJqE0")

# --- CONFIGURACIÓN DE GOOGLE SERVICES ---
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

client = None
sheet = None
drive_service = None

def inicializar_servicios_google():
    global client, sheet, drive_service
    try:
        # Intenta cargar desde archivo o variables de entorno
        if os.path.exists('credenciales.json'):
            creds = ServiceAccountCredentials.from_json_keyfile_name('credenciales.json', scope)
        else:
            # Alternativa si se configuran variables de entorno en Render
            creds = ServiceAccountCredentials.from_json_keyfile_name('/etc/secrets/credenciales.json', scope)
        
        client = gspread.authorize(creds)
        sheet = client.open("Base_Datos_Calculadora").sheet1
        drive_service = build('drive', 'v3', credentials=creds)
        print(" Conexión exitosa a Google Sheets y Google Drive API.")
    except Exception as e:
        print(f" Error de conexión a Google Services: {e}")

inicializar_servicios_google()

# --- ESTRUCTURA DE CACHÉ INTERNA PARA METAS ---
pdf_metas_cache = {
    "estilos": ["ESTILO-A", "ESTILO-B", "ESTILO-C", "ESTILO-D"], 
    "tallas": ['XXS', 'XS', 'S', 'M', 'L', 'XL', '2X', '3X', '4X'], 
    "procesos": ['CONTEO', 'SORTEO', 'VOLTEO', 'DOBLADO', 'VOLTEO-SORTING', 'VOLTEO-PFD', 'SORTEO-REPROCESO'], 
    "datos": [
        {"estilo": "ESTILO-A", "talla": "M", "proceso": "DOBLADO", "meta": 50},
        {"estilo": "ESTILO-B", "talla": "L", "proceso": "SORTEO", "meta": 65},
        {"estilo": "ESTILO-C", "talla": "S", "proceso": "VOLTEO", "meta": 80}
    ]
}

# --- FUNCIONES AUXILIARES PARA GOOGLE DRIVE Y ARCHIVOS ---

def obtener_o_crear_carpeta_usuario(nombre_usuario):
    if not drive_service:
        return None
    try:
        nombre_sanitizado = str(nombre_usuario).strip().replace("'", "\\'")
        query = f"'{CARPETA_RAIZ_DRIVE}' in parents and name = '{nombre_sanitizado}' and mimeType = 'application/vnd.google-apps.folder' and trashed = false"
        results = drive_service.files().list(q=query, fields="files(id, name)").execute()
        files = results.get('files', [])

        if files:
            return files[0]['id']

        file_metadata = {
            'name': nombre_usuario,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [CARPETA_RAIZ_DRIVE]
        }
        folder = drive_service.files().create(body=file_metadata, fields='id').execute()
        return folder.get('id')
    except Exception as e:
        print(f"Error al gestionar carpeta del usuario {nombre_usuario}: {e}")
        return None

def generar_nombre_correlativo(folder_id):
    if not drive_service or not folder_id:
        fecha_actual = datetime.now().strftime("%d-%m-%Y")
        return f"calculo_000001_{fecha_actual}"
    try:
        query = f"'{folder_id}' in parents and trashed = false"
        results = drive_service.files().list(q=query, fields="files(name)").execute()
        files = results.get('files', [])

        numero_calculo = len(files) + 1
        str_numero = f"{numero_calculo:06d}"
        fecha_actual = datetime.now().strftime("%d-%m-%Y")

        return f"calculo_{str_numero}_{fecha_actual}"
    except Exception as e:
        print(f"Error al generar correlativo: {e}")
        fecha_actual = datetime.now().strftime("%d-%m-%Y")
        return f"calculo_000001_{fecha_actual}"

def crear_pdf_en_memoria(titulo_modulo, datos_extensos):
    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)
    
    # Encabezado principal
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, 750, f"VEKTOR NEXUS - REPORTE DE {titulo_modulo.upper()}")
    c.setFont("Helvetica", 10)
    c.drawString(50, 735, f"Fecha y Hora de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    c.line(50, 725, 550, 725)

    y = 695
    c.setFont("Helvetica", 11)
    
    if isinstance(datos_extensos, list):
        for linea in datos_extensos:
            if y < 50:
                c.showPage()
                c.setFont("Helvetica", 11)
                y = 750
            c.drawString(50, y, str(linea))
            y -= 22
    elif isinstance(datos_extensos, dict):
        for k, v in datos_extensos.items():
            if y < 50:
                c.showPage()
                c.setFont("Helvetica", 11)
                y = 750
            c.drawString(50, y, f"{k}: {v}")
            y -= 22
    else:
        for linea in str(datos_extensos).split('|'):
            if y < 50:
                c.showPage()
                c.setFont("Helvetica", 11)
                y = 750
            c.drawString(50, y, linea.strip())
            y -= 22

    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer

def crear_imagen_en_memoria(titulo, datos_cortos):
    img = Image.new('RGB', (650, 350), color='#050814')
    d = ImageDraw.Draw(img)

    d.text((30, 25), f"VEKTOR NEXUS: {titulo.upper()}", fill='#00f3ff')
    d.line([(30, 55), (620, 55)], fill='#ffe600', width=2)

    y = 75
    if isinstance(datos_cortos, dict):
        for k, v in datos_cortos.items():
            d.text((30, y), f"{k}: {v}", fill='#edf2f4')
            y += 35
    else:
        d.text((30, y), f"Detalle: {str(datos_cortos)}", fill='#edf2f4')

    img_buffer = io.BytesIO()
    img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    return img_buffer

def cargar_usuarios_drive():
    if not sheet:
        return {}
    try:
        records = sheet.get_all_values()
        usuarios = {}
        for row in records[1:]:
            if len(row) > 0 and str(row[0]).strip():
                tkn = str(row[0]).strip()
                # Columna 6 (índice 6) define estado de hibernación/bloqueo
                is_hibernated = str(row[6]).lower() == 'true' if len(row) > 6 and row[6] != "" else False

                usuarios[tkn] = {
                    "token": tkn,
                    "nombre": str(row[1]).strip() if len(row) > 1 else "Operador",
                    "contacto": str(row[2]).strip() if len(row) > 2 else "",
                    "pin": str(row[3]).strip() if len(row) > 3 else "0000",
                    "rol": str(row[4]).strip() if len(row) > 4 else "operador",
                    "device_id": str(row[5]).strip() if len(row) > 5 else "",
                    "hibernado": is_hibernated,
                    "ultima_conexion": str(row[11]).strip() if len(row) > 11 else "Desconocida",
                    "permisos": {
                        "biohorario": str(row[6]).lower() == 'true' if len(row) > 6 and row[6] != "" else True,
                        "eficiencia": str(row[7]).lower() == 'true' if len(row) > 7 and row[7] != "" else True,
                        "tiempo": str(row[8]).lower() == 'true' if len(row) > 8 and row[8] != "" else True,
                        "metas": str(row[9]).lower() == 'true' if len(row) > 9 and row[9] != "" else True,
                        "historial": str(row[10]).lower() == 'true' if len(row) > 10 and row[10] != "" else True
                    }
                }
        return usuarios
    except Exception as e:
        print("Error al cargar lista de usuarios desde Google Sheets:", e)
        return {}

def procesar_archivos_excel_drive():
    """Busca archivos .xlsx en Google Drive e integra sus celdas al caché de metas."""
    if not drive_service:
        return pdf_metas_cache

    try:
        query = f"'{CARPETA_RAIZ_DRIVE}' in parents and mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed = false"
        results = drive_service.files().list(q=query, fields="files(id, name)").execute()
        excel_files = results.get('files', [])

        nuevos_estilos = set(pdf_metas_cache["estilos"])
        nuevos_procesos = set(pdf_metas_cache["procesos"])
        nuevos_datos = list(pdf_metas_cache["datos"])

        for f in excel_files:
            file_id = f['id']
            request_file = drive_service.files().get_media(fileId=file_id)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request_file)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            
            fh.seek(0)
            wb = openpyxl.load_workbook(fh, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) >= 4 and row[0] and row[1] and row[2] and row[3]:
                        estilo = str(row[0]).strip().upper()
                        talla = str(row[1]).strip().upper()
                        proceso = str(row[2]).strip().upper()
                        try:
                            meta_val = float(row[3])
                        except:
                            continue

                        nuevos_estilos.add(estilo)
                        nuevos_procesos.add(proceso)
                        
                        # Actualizar o insertar meta
                        idx = next((i for i, item in enumerate(nuevos_datos) if item["estilo"] == estilo and item["talla"] == talla and item["proceso"] == proceso), -1)
                        if idx >= 0:
                            nuevos_datos[idx]["meta"] = meta_val
                        else:
                            nuevos_datos.append({"estilo": estilo, "talla": talla, "proceso": proceso, "meta": meta_val})

        pdf_metas_cache["estilos"] = sorted(list(nuevos_estilos))
        pdf_metas_cache["procesos"] = sorted(list(nuevos_procesos))
        pdf_metas_cache["datos"] = nuevos_datos

    except Exception as e:
        print(f"Error al procesar archivos Excel de Drive: {e}")

    return pdf_metas_cache

# --- RUTAS PRINCIPALES Y ENDPOINTS DE API ---

@app.route('/')
def index():
    token = request.args.get('token', '')
    usuarios_actuales = cargar_usuarios_drive()
    
    # Fallback para desarrollo si no existen datos en Google Sheets aún
    if token not in usuarios_actuales and token in ['angel0301', 'libny534']:
        user_data = {
            "token": token,
            "nombre": "Angel Castillo",
            "contacto": "+50499999999",
            "pin": "0301",
            "rol": "admin",
            "hibernado": False,
            "permisos": {"biohorario": True, "eficiencia": True, "tiempo": True, "metas": True, "historial": True}
        }
    elif token in usuarios_actuales:
        user_data = usuarios_actuales[token]
    else:
        return "<div style='background:#050814;color:#ff007f;text-align:center;padding:100px;font-family:sans-serif;height:100vh;'><h1>ACCESO DENEGADO</h1><p>Token de seguridad no reconocido o caducado.</p></div>", 403

    return render_template('index.html', user=user_data, token=token)

@app.route('/api/login', methods=['POST'])
def login_verificar():
    data = request.json or {}
    token = data.get('token')
    pin_ingresado = str(data.get('pin', '')).strip()

    usuarios_actuales = cargar_usuarios_drive()
    
    # Soporte para cuentas maestras por defecto
    if token in ['angel0301', 'libny534'] and token not in usuarios_actuales:
        usuarios_actuales[token] = {
            "pin": "0301" if token == 'angel0301' else "534",
            "nombre": "Angel Castillo",
            "hibernado": False,
            "permisos": {"biohorario": True, "eficiencia": True, "tiempo": True, "metas": True, "historial": True}
        }

    if token in usuarios_actuales and str(usuarios_actuales[token]['pin']).strip() == pin_ingresado:
        user = usuarios_actuales[token]
        session['user_token'] = token
        session['user_name'] = user.get('nombre', 'Operador')
        
        return jsonify({
            "status": "success",
            "hibernacion": user.get('hibernado', False),
            "permisos": user.get('permisos', {})
        })
        
    return jsonify({"status": "error", "message": "PIN o Token de acceso incorrecto"}), 401

@app.route('/api/metas/datos', methods=['GET'])
def obtener_metas_datos():
    return jsonify({
        "status": "success",
        "datos": pdf_metas_cache["datos"],
        "estilos": pdf_metas_cache["estilos"],
        "tallas": pdf_metas_cache["tallas"],
        "procesos": pdf_metas_cache["procesos"]
    })

@app.route('/api/metas/sincronizar', methods=['POST'])
def sincronizar_metas():
    metas_actualizadas = procesar_archivos_excel_drive()
    return jsonify({
        "status": "success",
        "datos": metas_actualizadas["datos"],
        "estilos": metas_actualizadas["estilos"],
        "tallas": metas_actualizadas["tallas"],
        "procesos": metas_actualizadas["procesos"]
    })

# Mapeo dual para compatibilidad con las solicitudes /api/save y /api/historial/guardar
@app.route('/api/save', methods=['POST'])
@app.route('/api/historial/guardar', methods=['POST'])
def guardar_calculo():
    data = request.json or {}
    token = data.get('token') or session.get('user_token')
    tipo_calculo = data.get('tipo', 'Cálculo General')
    info = data.get('info', {})
    extenso = data.get('extenso', False)
    lineas = data.get('lineas', [])

    if not token:
        return jsonify({"status": "error", "message": "Token de sesión faltante"}), 400

    usuarios = cargar_usuarios_drive()
    nombre_usuario = usuarios.get(token, {}).get('nombre', 'Operador_Anonimo')
    
    folder_id = obtener_o_crear_carpeta_usuario(nombre_usuario)
    nombre_archivo_base = generar_nombre_correlativo(folder_id)

    # Generación dinámica del reporte PDF o PNG según el tipo de datos
    if extenso or len(lineas) > 4:
        contenido_binario = crear_pdf_en_memoria(tipo_calculo, lineas if lineas else info)
        nombre_completo = f"{nombre_archivo_base}_{tipo_calculo.replace(' ', '_')}.pdf"
        mime = "application/pdf"
    else:
        contenido_binario = crear_imagen_en_memoria(tipo_calculo, info)
        nombre_completo = f"{nombre_archivo_base}_{tipo_calculo.replace(' ', '_')}.png"
        mime = "image/png"

    try:
        if drive_service and folder_id:
            file_metadata = {'name': nombre_completo, 'parents': [folder_id]}
            media = MediaIoBaseUpload(contenido_binario, mimetype=mime, resumable=True)
            uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
            
            drive_url = uploaded_file.get('webViewLink')
            file_id = uploaded_file.get('id')
        else:
            drive_url = "#"
            file_id = "local_dummy_id"

        return jsonify({
            "status": "success",
            "success": True,
            "message": "Cálculo registrado en Google Drive con éxito",
            "file_name": nombre_completo,
            "file_id": file_id,
            "drive_url": drive_url
        })
    except Exception as e:
        print(f"Error al subir cálculo a Google Drive: {e}")
        return jsonify({"status": "error", "message": f"Error al subir a Drive: {str(e)}"}), 500

# Mapeo dual para compatibilidad con las solicitudes /api/load y /api/historial/archivos
@app.route('/api/load', methods=['GET'])
@app.route('/api/historial/archivos', methods=['GET'])
def listar_historial_usuario():
    token = request.args.get('token') or session.get('user_token')
    if not token:
        return jsonify({"status": "error", "message": "Falta parámetro token"}), 400

    usuarios = cargar_usuarios_drive()
    nombre_usuario = usuarios.get(token, {}).get('nombre', 'Operador')
    folder_id = obtener_o_crear_carpeta_usuario(nombre_usuario)

    if not folder_id or not drive_service:
        return jsonify([])

    try:
        query = f"'{folder_id}' in parents and trashed = false"
        results = drive_service.files().list(q=query, fields="files(id, name, webViewLink, mimeType, createdTime)").execute()
        files = results.get('files', [])

        formateados = []
        for f in files:
            formateados.append({
                "tipo": f.get('name', 'Reporte').split('_')[1] if '_' in f.get('name', '') else "Reporte Generado",
                "fecha_hora": f.get('createdTime', datetime.now().strftime("%d/%m/%Y %H:%M")),
                "drive_url": f.get('webViewLink'),
                "usuario": nombre_usuario,
                "info": {
                    "res": f.get('name'),
                    "detalle": f"Tipo de documento: {f.get('mimeType')}"
                }
            })
        return jsonify(formateados)
    except Exception as e:
        print(f"Error al obtener historial de Drive: {e}")
        return jsonify([]), 500

@app.route('/api/admin/usuarios', methods=['GET', 'POST', 'PUT'])
def admin_drive():
    if request.method == 'GET':
        return jsonify(cargar_usuarios_drive())

    data = request.json or {}
    if request.method == 'POST':
        nuevo_token = "tkn_" + "".join(random.choices(string.ascii_lowercase + string.digits, k=6))
        nuevo_pin = "".join(random.choices(string.digits, k=4))
        try:
            if sheet:
                sheet.append_row([nuevo_token, data.get('nombre'), data.get('contacto'), nuevo_pin, "operador", "", "false", "true", "true", "true", "true", ""])
            return jsonify({"status": "success", "token": nuevo_token, "pin": nuevo_pin})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

    if request.method == 'PUT':
        token = data.get('token')
        if not sheet:
            return jsonify({"status": "error", "message": "No hay conexión con Google Sheets"}), 500
        try:
            celda = sheet.find(token)
            if data.get('nombre'): sheet.update_cell(celda.row, 2, data.get('nombre'))
            if data.get('contacto'): sheet.update_cell(celda.row, 3, data.get('contacto'))
            if data.get('nuevo_pin'): sheet.update_cell(celda.row, 4, data.get('nuevo_pin'))
            if 'hibernar' in data: sheet.update_cell(celda.row, 7, str(data.get('hibernar')).lower())
            return jsonify({"status": "success"})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 404

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
